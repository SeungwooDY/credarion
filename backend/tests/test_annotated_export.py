"""Tests for the annotated statement export and the date-tolerance window.

Covers:
  - _classify_review date-gap handling (match-with-note, never a discrepancy)
  - classify_result wording/fill for each result state
  - build_annotated_workbook in both in-place and rebuild modes
  - cleaning keeps ingestor helper columns out of raw_row
"""
from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.ingestion.cleaning import clean_dataframe
from app.reconciliation.annotate import (
    NOTE_HEADER,
    RESULT_HEADER,
    build_annotated_workbook,
    classify_result,
)
from app.reconciliation.exact_match import MatchCandidate, MatchResult, StatementItem
from app.reconciliation.orchestrator import _classify_review


# ============================================================
# Date tolerance (match-with-note)
# ============================================================


def _make_match(grn_date, delivery_date, qty_delta=Decimal("0")) -> MatchResult:
    erp = MatchCandidate(
        erp_id=1, po_number="428759", material_number="MAT001",
        quantity=Decimal("100"), po_price=Decimal("10"),
        amount=Decimal("1000"), grn_date=grn_date,
    )
    stmt = StatementItem(
        line_id=2, po_number="428759", material_number="MAT001",
        quantity=Decimal("100") + qty_delta, unit_price=Decimal("10"),
        amount=Decimal("1000"), delivery_date=delivery_date,
    )
    return MatchResult(
        erp=erp, statement=stmt, match_type="exact",
        quantity_delta=qty_delta, price_delta=Decimal("0"),
        amount_delta=Decimal("0"),
        status="discrepancy" if qty_delta else "matched",
    )


class TestDateToleranceClassification:
    def test_within_window_stays_exact(self):
        m = _make_match(datetime(2026, 3, 10), date(2026, 3, 12))
        review = _classify_review(m, date_tolerance_days=3)
        assert review["match_type"] == "exact"
        assert review["discrepancy_note"] is None

    def test_beyond_window_becomes_match_with_note(self):
        m = _make_match(datetime(2026, 3, 1), date(2026, 3, 20))
        review = _classify_review(m, date_tolerance_days=3)
        assert review["match_type"] == "near_exact"
        assert review["confidence_label"] == "Match — Date Mismatch"
        assert "19 days" in review["discrepancy_note"]
        # A date gap alone is NOT a discrepancy — status stays reviewable.
        assert review["status"] == "pending_review"

    def test_wider_tolerance_suppresses_note(self):
        m = _make_match(datetime(2026, 3, 1), date(2026, 3, 20))
        review = _classify_review(m, date_tolerance_days=30)
        assert review["match_type"] == "exact"
        assert review["discrepancy_note"] is None

    def test_missing_statement_date_ignored(self):
        m = _make_match(datetime(2026, 3, 1), None)
        review = _classify_review(m, date_tolerance_days=3)
        assert review["match_type"] == "exact"
        assert review["discrepancy_note"] is None

    def test_date_note_appended_to_existing_near_exact_note(self):
        m = _make_match(datetime(2026, 3, 1), date(2026, 3, 20), qty_delta=Decimal("5"))
        review = _classify_review(m, date_tolerance_days=3)
        assert review["match_type"] == "near_exact"
        # Both the qty-delta note and the date note are present.
        assert "Quantity" in review["discrepancy_note"]
        assert "19 days" in review["discrepancy_note"]


# ============================================================
# classify_result (annotation wording)
# ============================================================


def _result(**kw) -> SimpleNamespace:
    base = dict(
        status="pending_review", discrepancy_type=None,
        quantity_delta=None, price_delta=None, amount_delta=None,
        discrepancy_note=None, resolution_note=None,
    )
    base.update(kw)
    return SimpleNamespace(**base)


class TestClassifyResult:
    def test_clean_match(self):
        text, note, _ = classify_result(_result(status="confirmed"))
        assert text == "相符 ✓"
        assert note == ""

    def test_qty_and_price_discrepancy(self):
        text, note, _ = classify_result(_result(
            discrepancy_type="quantity_over,price_higher",
            quantity_delta=Decimal("5"), price_delta=Decimal("0.5"),
        ))
        assert text == "差异 ✗"
        assert "数量差异 +5" in note
        assert "单价差异 +0.5" in note

    def test_missing_from_erp(self):
        text, note, _ = classify_result(_result(discrepancy_type="missing_from_erp"))
        assert text == "差异 ✗"
        assert "ERP" in note

    def test_resolved_is_verified_with_note(self):
        text, note, _ = classify_result(_result(
            status="resolved", discrepancy_type="quantity_over",
            quantity_delta=Decimal("5"), resolution_note="供应商重复行，已合并",
        ))
        assert text == "已核实 ✓"
        assert note == "供应商重复行，已合并"

    def test_rejected_reason_included(self):
        _, note, _ = classify_result(_result(
            status="rejected", discrepancy_type="quantity_over",
            quantity_delta=Decimal("-3"), discrepancy_note="Supplier over-billed",
        ))
        assert "数量差异 -3" in note
        assert "Supplier over-billed" in note

    def test_fine_decimals_never_rounded(self):
        # Suppliers quote prices to the 4th decimal place and beyond
        # (e.g. 0.6575) — deltas must render at full precision.
        _, note, _ = classify_result(_result(
            discrepancy_type="price_higher",
            price_delta=Decimal("0.6575"),
        ))
        assert "单价差异 +0.6575" in note

        _, note, _ = classify_result(_result(
            discrepancy_type="price_higher",
            price_delta=Decimal("0.123456"),
        ))
        assert "单价差异 +0.123456" in note

    def test_trailing_zeros_trimmed_not_rounded(self):
        _, note, _ = classify_result(_result(
            discrepancy_type="price_higher",
            price_delta=Decimal("0.650000"),
        ))
        assert "单价差异 +0.65" in note


# ============================================================
# build_annotated_workbook
# ============================================================


def _line(source_row, raw_row=None):
    return SimpleNamespace(
        id=uuid.uuid4(), source_row=source_row, raw_row=raw_row or {},
    )


def _missing_erp():
    return SimpleNamespace(
        po_number="428800", material_number="590*5166*8*003",
        grn_number="GRN-77", grn_date=datetime(2026, 3, 28),
        quantity=Decimal("40"), po_price=Decimal("2.5"), amount=Decimal("100"),
    )


class TestBuildAnnotatedWorkbook:
    def test_in_place_annotation(self, tmp_path):
        # Original supplier file: header row 1, two data rows.
        path = tmp_path / "statement.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["订单号", "物料编码", "数量", "金额"])
        ws.append(["428759", "MAT001", 100, 1000])
        ws.append(["428760", "MAT002", 50, 500])
        wb.save(path)

        lines = [_line(2), _line(3)]
        results = {
            lines[0].id: _result(status="confirmed"),
            lines[1].id: _result(
                discrepancy_type="quantity_over", quantity_delta=Decimal("5"),
            ),
        }
        buf = build_annotated_workbook(
            file_path=str(path), header_excel_row=1,
            line_items=lines, result_by_line=results,
            missing_erp=[_missing_erp()],
        )

        out = load_workbook(buf).active
        # Supplier's own cells are untouched.
        assert out.cell(row=2, column=1).value == "428759"
        # Annotation columns appended after the original 4.
        assert out.cell(row=1, column=5).value == RESULT_HEADER
        assert out.cell(row=1, column=6).value == NOTE_HEADER
        assert out.cell(row=2, column=5).value == "相符 ✓"
        assert out.cell(row=3, column=5).value == "差异 ✗"
        assert "数量差异 +5" in out.cell(row=3, column=6).value
        # Missing-from-statement appendix present with the ERP receipt.
        found = [
            c.value for row in out.iter_rows() for c in row
            if c.value == "590*5166*8*003"
        ]
        assert found

    def test_rebuild_mode_when_file_missing(self):
        lines = [
            _line(6, {"订单号": "428759", "物料编码": "MAT001", "数量": "100"}),
            _line(7, {"订单号": "428760", "物料编码": "MAT002", "数量": "50"}),
        ]
        results = {lines[0].id: _result(status="confirmed")}
        buf = build_annotated_workbook(
            file_path="/nonexistent/original.xlsx", header_excel_row=5,
            line_items=lines, result_by_line=results, missing_erp=[],
        )
        out = load_workbook(buf).active
        # Rebuilt sheet: header row 1 from raw_row keys, data preserves order.
        assert out.cell(row=1, column=1).value == "订单号"
        assert out.cell(row=2, column=1).value == "428759"
        assert out.cell(row=1, column=4).value == RESULT_HEADER
        assert out.cell(row=2, column=4).value == "相符 ✓"
        # Second line had no result in the run — annotation cell left empty.
        assert out.cell(row=3, column=4).value is None


# ============================================================
# Cleaning: helper columns stay out of raw_row
# ============================================================


class TestSourceRowHelperColumn:
    def test_source_row_survives_cleaning_and_stays_out_of_raw_row(self):
        df = pd.DataFrame({
            "订单号": ["428759", "合计", "428760"],
            "数量": ["100", "150", "50"],
        })
        df["_source_row"] = df.index + 7  # as the ingestor would
        cleaned = clean_dataframe(df, {"po_number": "订单号", "quantity": "数量"})
        # Summary row dropped; source rows preserved for the survivors.
        assert list(cleaned["_source_row"]) == [7, 9]
        for raw in cleaned["_raw_row"]:
            assert "_source_row" not in raw
            assert "订单号" in raw
