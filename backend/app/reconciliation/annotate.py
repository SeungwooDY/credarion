"""Annotated statement export.

Reproduces the supplier's ORIGINAL statement Excel and writes the
reconciliation outcome back onto it — mirroring the accountant's manual
workflow of marking discrepancies on the statement and sending it back to
the supplier.

Two annotation columns are appended to the right of the supplier's own
columns: 对账结果 (result) and 备注 (note). ERP receipts that are missing
from the statement are appended as an extra section below the data so the
supplier can include them next month.

Modes:
  - In-place: the original .xlsx is still on disk and line items carry
    source_row → the supplier's exact layout is preserved.
  - Rebuild: original file missing (pre-durable-storage uploads) or not
    .xlsx → the sheet is reconstructed from each line's raw_row audit copy.
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

RESULT_HEADER = "对账结果"
NOTE_HEADER = "备注"

_FILL_MATCHED = PatternFill("solid", start_color="E7F6EC")  # light green
_FILL_DISCREPANCY = PatternFill("solid", start_color="FFF2CC")  # amber
_FILL_MISSING = PatternFill("solid", start_color="FDE2E2")  # light red


def _fmt_delta(val: Any) -> str:
    """Signed delta at full precision — supplier decimals are never rounded
    (e.g. 0.6575 stays 0.6575); only trailing zeros are trimmed."""
    from decimal import Decimal

    d = val if isinstance(val, Decimal) else Decimal(str(val))
    s = format(d, "+f")
    return s.rstrip("0").rstrip(".") if "." in s else s


def classify_result(r: Any) -> tuple[str, str, PatternFill]:
    """(result text, note, fill) for one ReconciliationResult.

    Wording is Chinese-first since the annotated file goes back to the
    supplier's finance team. "resolved" means the discrepancy was reviewed
    and dismissed as false, so it renders as verified-with-note.
    """
    if r.status == "resolved":
        return "已核实 ✓", r.resolution_note or "", _FILL_MATCHED

    if r.discrepancy_type == "missing_from_erp":
        note = "ERP系统中未找到对应入库记录 (no matching ERP receipt)"
        if r.status == "rejected" and r.discrepancy_note:
            note = f"{note}；{r.discrepancy_note}"
        return "差异 ✗", note, _FILL_MISSING

    if r.discrepancy_type:
        parts: list[str] = []
        if r.quantity_delta is not None and float(r.quantity_delta) != 0:
            parts.append(f"数量差异 {_fmt_delta(r.quantity_delta)}")
        if r.price_delta is not None and float(r.price_delta) != 0:
            parts.append(f"单价差异 {_fmt_delta(r.price_delta)}")
        if not parts and r.amount_delta is not None and float(r.amount_delta) != 0:
            parts.append(f"金额差异 {_fmt_delta(r.amount_delta)}")
        note = "；".join(parts)
        # A human-entered reason (flag/reject) takes precedence over engine text.
        if r.status == "rejected" and r.discrepancy_note:
            note = f"{note}；{r.discrepancy_note}" if note else r.discrepancy_note
        return "差异 ✗", note, _FILL_DISCREPANCY

    if r.status == "rejected":
        return "差异 ✗", r.discrepancy_note or "", _FILL_DISCREPANCY

    return "相符 ✓", "", _FILL_MATCHED


def _can_annotate_in_place(file_path: str | None, line_items: list[Any]) -> bool:
    if not file_path or Path(file_path).suffix.lower() != ".xlsx":
        return False
    if not Path(file_path).exists():
        return False
    return all(li.source_row is not None for li in line_items)


def _rebuild_sheet(ws: Any, line_items: list[Any]) -> tuple[dict[Any, int], int]:
    """Reconstruct the statement from raw_row copies.

    Returns (line_id → sheet row, header row). Column order follows the
    original file because raw_row dicts preserve insertion order.
    """
    columns: list[str] = []
    for li in line_items:
        for col in (li.raw_row or {}):
            if col not in columns:
                columns.append(col)
    if not columns:
        columns = ["订单号", "物料编码", "数量", "单价", "金额"]

    for idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=col).font = Font(bold=True)

    row_for_line: dict[Any, int] = {}
    ordered = sorted(line_items, key=lambda li: li.source_row or 0)
    for row_idx, li in enumerate(ordered, start=2):
        raw = li.raw_row or {}
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=raw.get(col))
        row_for_line[li.id] = row_idx
    return row_for_line, 1


def _append_missing_section(ws: Any, missing_erp: list[Any]) -> None:
    """ERP receipts absent from the statement → extra rows below the data."""
    if not missing_erp:
        return
    start = ws.max_row + 2
    title = ws.cell(
        row=start,
        column=1,
        value="以下项目ERP有入库记录，但本对账单中缺失 (in ERP but missing from this statement):",
    )
    title.font = Font(bold=True)
    title.fill = _FILL_MISSING

    headers = ["订单号", "物料编码", "入库单号", "入库日期", "数量", "单价", "金额"]
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=start + 1, column=idx, value=h)
        c.font = Font(bold=True)
        c.fill = _FILL_MISSING

    for offset, erp in enumerate(missing_erp, start=2):
        grn_date = erp.grn_date
        if isinstance(grn_date, datetime):
            grn_date = grn_date.date()
        values = [
            erp.po_number,
            erp.material_number,
            erp.grn_number,
            str(grn_date) if grn_date else None,
            float(erp.quantity) if erp.quantity is not None else None,
            float(erp.po_price) if erp.po_price is not None else None,
            float(erp.amount) if erp.amount is not None else None,
        ]
        for idx, v in enumerate(values, start=1):
            ws.cell(row=start + offset, column=idx, value=v).fill = _FILL_MISSING


def build_annotated_workbook(
    file_path: str | None,
    header_excel_row: int | None,
    line_items: list[Any],
    result_by_line: dict[Any, Any],
    missing_erp: list[Any],
) -> io.BytesIO:
    """Produce the annotated statement as an in-memory .xlsx.

    Args:
        file_path: stored path of the original upload (may be gone).
        header_excel_row: 1-based row holding the supplier's column headers
            (falls back to min(source_row) - 1, then row 1).
        line_items: the statement's StatementLineItem rows.
        result_by_line: line_id → ReconciliationResult (latest run).
        missing_erp: ERPRecord rows flagged missing_from_statement.
    """
    if _can_annotate_in_place(file_path, line_items):
        wb = load_workbook(file_path)
        ws = wb.active
        row_for_line = {li.id: li.source_row for li in line_items}
        if header_excel_row is None:
            rows = [li.source_row for li in line_items if li.source_row]
            header_excel_row = min(rows) - 1 if rows else 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "对账单"
        row_for_line, header_excel_row = _rebuild_sheet(ws, line_items)

    col_result = ws.max_column + 1
    col_note = col_result + 1
    ws.cell(row=header_excel_row, column=col_result, value=RESULT_HEADER).font = Font(bold=True)
    ws.cell(row=header_excel_row, column=col_note, value=NOTE_HEADER).font = Font(bold=True)
    ws.column_dimensions[ws.cell(row=1, column=col_note).column_letter].width = 50

    for li in line_items:
        r = result_by_line.get(li.id)
        row = row_for_line.get(li.id)
        if r is None or row is None:
            continue
        text, note, fill = classify_result(r)
        cell = ws.cell(row=row, column=col_result, value=text)
        cell.fill = fill
        if note:
            ws.cell(row=row, column=col_note, value=note)

    _append_missing_section(ws, missing_erp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
